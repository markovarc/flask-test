import sqlite3
from flask import Flask, request, redirect, send_file
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'supersecretkey123'
app.config['DATABASE'] = 'an30.db'
app.config['SQLITE_TIMEOUT'] = 20

COLORS = {
    'primary': "#6C7A89",
    'secondary': "#95A5A6",
    'background': "#F5F7FA",
    'accent': "#4A90E2",
    'danger': "#ff4444",
    'status': {
        'work': "#C8E6C9",
        'stop': "#FFCDD2",
        'repair': "#FFF9C4",
        'holiday': "#E1BEE7"
    }
}

def init_db():
    with app.app_context():
        conn = sqlite3.connect(app.config['DATABASE'], timeout=app.config['SQLITE_TIMEOUT'])
        conn.execute("PRAGMA foreign_keys = ON")
        c = conn.cursor()

        # ОЧИСТКА (УДАЛЕНИЕ) ВСЕХ ТАБЛИЦ, ЧТОБЫ НЕ БЫЛО КОНФЛИКТОВ С УЖЕ СУЩЕСТВУЮЩИМИ ТАБЛИЦАМИ
        c.execute("DROP TABLE IF EXISTS records")
        c.execute("DROP TABLE IF EXISTS machines")
        c.execute("DROP TABLE IF EXISTS drivers")
        c.execute("DROP TABLE IF EXISTS counterparties")

        # Таблица техники
        c.execute('''CREATE TABLE IF NOT EXISTS machines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )''')

        # Таблица водителей
        c.execute('''CREATE TABLE IF NOT EXISTS drivers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )''')

        # Таблица контрагентов
        c.execute('''CREATE TABLE IF NOT EXISTS counterparties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )''')

        # Основная таблица записей
        c.execute('''CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date DATE NOT NULL,
            machine_id INTEGER NOT NULL,
            driver_id INTEGER NOT NULL,
            start_time TEXT,
            end_time TEXT,
            hours INTEGER DEFAULT 0,
            comment TEXT,
            counterparty_id INTEGER,
            status TEXT NOT NULL CHECK(status IN ('work', 'stop', 'repair', 'holiday')),
            FOREIGN KEY(machine_id) REFERENCES machines(id) ON DELETE CASCADE,
            FOREIGN KEY(driver_id) REFERENCES drivers(id) ON DELETE CASCADE,
            FOREIGN KEY(counterparty_id) REFERENCES counterparties(id) ON DELETE SET NULL
        )''')

        conn.commit()
        conn.close()

def get_db():
    conn = sqlite3.connect(app.config['DATABASE'], timeout=app.config['SQLITE_TIMEOUT'])
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def render_base(content):
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>АН-30 Учёт</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        body {{
            font-family: 'Segoe UI', sans-serif;
            background: {COLORS['background']};
            color: {COLORS['primary']};
        }}
        .header {{
            background: {COLORS['primary']};
            padding: 1rem;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        .nav {{
            max-width: 1200px;
            margin: 0 auto;
            display: flex;
            gap: 1rem;
        }}
        .nav a {{
            color: white;
            text-decoration: none;
            padding: 0.5rem 1rem;
            border-radius: 4px;
            transition: 0.3s;
        }}
        .nav a:hover {{
            background: {COLORS['secondary']};
        }}
        .container {{
            max-width: 1200px;
            margin: 2rem auto;
            padding: 0 1rem;
        }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 1.5rem;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            margin-bottom: 1rem;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }}
        th, td {{
            padding: 1rem;
            text-align: left;
            border-bottom: 1px solid #eee;
        }}
        th {{
            background: {COLORS['primary']};
            color: white;
        }}
        .status {{
            display: inline-block;
            padding: 0.25rem 0.75rem;
            border-radius: 1rem;
            font-size: 0.9em;
        }}
        .btn {{
            background: {COLORS['accent']};
            color: white;
            padding: 0.5rem 1rem;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            transition: 0.3s;
        }}
        .btn-danger {{
            background: {COLORS['danger']} !important;
        }}
        .btn:hover {{
            opacity: 0.9;
        }}
        .back-btn {{
            background: {COLORS['secondary']};
            margin: 1rem 0;
        }}
        form {{
            display: flex;
            gap: 1rem;
            flex-wrap: wrap;
        }}
        input, select {{
            padding: 0.5rem;
            border: 1px solid #ddd;
            border-radius: 4px;
            min-width: 250px;
        }}
        .calendar-grid {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 0.5rem;
        }}
        .calendar-day {{
            background: white;
            padding: 1rem;
            border-radius: 8px;
            min-height: 120px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
    </style>
</head>
<body>
    <header class="header">
        <nav class="nav">
            <a href="/">Главная</a>
            <a href="/admin">Админка</a>
            <a href="/export">📊 Отчёт</a>
        </nav>
    </header>
    <div class="container">
        {content}
    </div>
    <script>
        function confirmDelete(msg) {{
            return confirm(msg || 'Вы уверены что хотите удалить запись?');
        }}
    </script>
</body>
</html>'''

@app.route('/')
def index():
    conn = get_db()
    try:
        machines = conn.execute('SELECT * FROM machines').fetchall()
    finally:
        conn.close()
    return render_base(f'''
        <div class="card">
            <h1>Учёт работы спецтехники</h1>
            <table>
                <tr><th>Техника</th><th>Действия</th></tr>
                {''.join(f'''
                <tr>
                    <td>{row[1]}</td>
                    <td><a class="btn" href="/calendar/{row[0]}">📅 Календарь</a></td>
                </tr>
                ''' for row in machines)}
            </table>
        </div>
    ''')

@app.route('/calendar/<int:machine_id>')
def calendar(machine_id):
    conn = get_db()
    try:
        machine = conn.execute('SELECT * FROM machines WHERE id = ?', (machine_id,)).fetchone()
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        dates = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

        records = {}
        for d in dates:
            day_records = conn.execute('''
                SELECT drivers.name, status, start_time, end_time, counterparties.name
                FROM records
                JOIN drivers ON records.driver_id = drivers.id
                LEFT JOIN counterparties ON records.counterparty_id = counterparties.id
                WHERE machine_id = ? AND date = ?
            ''', (machine_id, d.date())).fetchall()
            records[d.date()] = day_records
    finally:
        conn.close()

    calendar_html = '<div class="calendar-grid">'
    for d in dates:
        day_records = records.get(d.date(), [])
        calendar_html += f'''
        <div class="calendar-day">
            <div style="font-weight: bold; margin-bottom: 0.5rem;">{d.strftime("%d.%m")}</div>
            {''.join(f'''
            <div class="status" style="background: {COLORS['status'].get(r[1], '#ffffff')}">
                {r[0]} - {r[1].capitalize()}<br>
                {f"{r[2]}-{r[3]}" if r[2] and r[3] else ""}<br>
                {r[4] if r[4] else ""}
            </div>
            ''' for r in day_records)}
        </div>'''
    calendar_html += '</div>'

    return render_base(f'''
        <a href="/" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>{machine[1]} - {today.strftime("%B %Y")}</h1>
            {calendar_html}
        </div>
    ''')

@app.route('/admin')
def admin():
    return render_base('''
        <div class="card">
            <h1>Административная панель</h1>
            <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem;">
                <a class="btn" href="/admin/machines">🚜 Управление техникой</a>
                <a class="btn" href="/admin/drivers">👤 Управление водителями</a>
                <a class="btn" href="/admin/counterparties">🏢 Управление контрагентами</a>
                <a class="btn" href="/admin/records">📅 Управление записями</a>
            </div>
        </div>
    ''')

@app.route('/admin/machines', methods=['GET', 'POST'])
def admin_machines():
    if request.method == 'POST':
        name = request.form['name']
        conn = get_db()
        try:
            conn.execute('INSERT INTO machines (name) VALUES (?)', (name,))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
        finally:
            conn.close()
        return redirect('/admin/machines')
    
    conn = get_db()
    try:
        machines = conn.execute('SELECT * FROM machines').fetchall()
    finally:
        conn.close()
    
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление техникой</h1>
            <form method="POST">
                <input type="text" name="name" placeholder="Название техники" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Название</th><th>Действия</th></tr>
                {''.join(f'''
                <tr>
                    <td>{row[0]}</td>
                    <td>{row[1]}</td>
                    <td>
                        <form method="POST" action="/delete/machine/{row[0]}">
                            <button type="submit" class="btn btn-danger" 
                                onclick="return confirmDelete('Удалить машину {row[1]}?')">
                                Удалить
                            </button>
                        </form>
                    </td>
                </tr>
                ''' for row in machines)}
            </table>
        </div>
    ''')

@app.route('/admin/drivers', methods=['GET', 'POST'])
def admin_drivers():
    if request.method == 'POST':
        name = request.form['name']
        conn = get_db()
        try:
            conn.execute('INSERT INTO drivers (name) VALUES (?)', (name,))
            conn.commit()
        except sqlite3.IntegrityError as e:
            print(f"Ошибка добавления водителя: {e}")
        finally:
            conn.close()
        return redirect('/admin/drivers')
    
    conn = get_db()
    try:
        drivers = conn.execute('SELECT * FROM drivers').fetchall()
    finally:
        conn.close()
    
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление водителями</h1>
            <form method="POST">
                <input type="text" name="name" placeholder="ФИО водителя" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Имя</th><th>Действия</th></tr>
                {''.join(f'''
                <tr>
                    <td>{row[0]}</td>
                    <td>{row[1]}</td>
                    <td>
                        <form method="POST" action="/delete/driver/{row[0]}">
                            <button type="submit" class="btn btn-danger"
                                onclick="return confirmDelete('Удалить водителя {row[1]}?')">
                                Удалить
                            </button>
                        </form>
                    </td>
                </tr>
                ''' for row in drivers)}
            </table>
        </div>
    ''')

@app.route('/admin/counterparties', methods=['GET', 'POST'])
def admin_counterparties():
    if request.method == 'POST':
        name = request.form['name']
        conn = get_db()
        try:
            conn.execute('INSERT INTO counterparties (name) VALUES (?)', (name,))
            conn.commit()
        except sqlite3.IntegrityError as e:
            print(f"Ошибка добавления контрагента: {e}")
        finally:
            conn.close()
        return redirect('/admin/counterparties')
    
    conn = get_db()
    try:
        counterparties = conn.execute('SELECT * FROM counterparties').fetchall()
    finally:
        conn.close()
    
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление контрагентами</h1>
            <form method="POST">
                <input type="text" name="name" placeholder="Название контрагента" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Название</th><th>Действия</th></tr>
                {''.join(f'''
                <tr>
                    <td>{row[0]}</td>
                    <td>{row[1]}</td>
                    <td>
                        <form method="POST" action="/delete/counterparty/{row[0]}">
                            <button type="submit" class="btn btn-danger"
                                onclick="return confirmDelete('Удалить контрагента {row[1]}?')">
                                Удалить
                            </button>
                        </form>
                    </td>
                </tr>
                ''' for row in counterparties)}
            </table>
        </div>
    ''')

@app.route('/admin/records', methods=['GET', 'POST'])
def admin_records():
    if request.method == 'POST':
        conn = get_db()
        try:
            date_str = request.form['date']
            machine_id = int(request.form['machine_id'])
            driver_id = int(request.form['driver_id'])
            status = request.form['status']
            start_time = request.form.get('start_time', '')
            end_time = request.form.get('end_time', '')
            comment = request.form.get('comment', '')

            counterparty_id = request.form.get('counterparty_id')
            if counterparty_id:
                counterparty_id = int(counterparty_id)
            else:
                counterparty_id = None

            hours = 0
            if start_time and end_time:
                try:
                    start = datetime.strptime(start_time, '%H:%M')
                    end = datetime.strptime(end_time, '%H:%M')
                    if end < start:
                        end += timedelta(days=1)
                    delta = end - start
                    hours = delta.seconds // 3600
                except ValueError as e:
                    print(f"Ошибка расчета времени: {e}")

            conn.execute('''
                INSERT INTO records
                (date, machine_id, driver_id, status, start_time, end_time, hours, comment, counterparty_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                date_str,
                machine_id,
                driver_id,
                status,
                start_time or None,
                end_time or None,
                hours,
                comment,
                counterparty_id
            ))
            conn.commit()
        except Exception as e:
            print(f"Ошибка создания записи: {e}")
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/records')

    conn = get_db()
    try:
        records = conn.execute('''
            SELECT r.id, r.date, m.name, d.name, r.start_time, r.end_time,
                   r.hours, r.comment, c.name, r.status
            FROM records r
            JOIN machines m ON r.machine_id = m.id
            JOIN drivers d ON r.driver_id = d.id
            LEFT JOIN counterparties c ON r.counterparty_id = c.id
            ORDER BY r.date DESC
        ''').fetchall()

        machines = conn.execute('SELECT * FROM machines').fetchall()
        drivers = conn.execute('SELECT * FROM drivers').fetchall()
        counterparties = conn.execute('SELECT * FROM counterparties').fetchall()
    finally:
        conn.close()

    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление записями</h1>
            <form method="POST">
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 1rem;">
                    <input type="date" name="date" required>
                    <select name="machine_id" required>
                        <option value="">Выберите технику</option>
                        {''.join(f'<option value="{row[0]}">{row[1]}</option>' for row in machines)}
                    </select>
                    <select name="driver_id" required>
                        <option value="">Выберите водителя</option>
                        {''.join(f'<option value="{row[0]}">{row[1]}</option>' for row in drivers)}
                    </select>
                    <select name="status" required>
                        <option value="work">Работа</option>
                        <option value="stop">Простой</option>
                        <option value="repair">Ремонт</option>
                        <option value="holiday">Выходной</option>
                    </select>
                    <input type="time" name="start_time" placeholder="Начало">
                    <input type="time" name="end_time" placeholder="Конец">
                    <select name="counterparty_id">
                        <option value="">Контрагент (не обязательно)</option>
                        {''.join(f'<option value="{row[0]}">{row[1]}</option>' for row in counterparties)}
                    </select>
                    <input type="text" name="comment" placeholder="Комментарий" style="grid-column: span 2;">
                </div>
                <button type="submit" class="btn" style="margin-top: 1rem;">Добавить запись</button>
            </form>

            <table style="margin-top: 2rem;">
                <tr>
                    <th>Дата</th>
                    <th>Техника</th>
                    <th>Водитель</th>
                    <th>Время</th>
                    <th>Часы</th>
                    <th>Контрагент</th>
                    <th>Комментарий</th>
                    <th>Статус</th>
                    <th>Действия</th>
                </tr>
                {''.join(f'''
                <tr>
                    <td>{datetime.strptime(row[1], '%Y-%m-%d').strftime('%d.%m.%Y')}</td>
                    <td>{row[2]}</td>
                    <td>{row[3]}</td>
                    <td>{f"{row[4]} - {row[5]}" if row[4] and row[5] else "-"}</td>
                    <td>{row[6] or "0"}</td>
                    <td>{row[8] or "-"}</td>
                    <td>{row[7] or "-"}</td>
                    <td>
                        <div class="status" style="background: {COLORS['status'].get(row[9], '#ffffff')}">
                            {row[9].capitalize()}
                        </div>
                    </td>
                    <td>
                        <form method="POST" action="/delete/record/{row[0]}">
                            <button type="submit" class="btn btn-danger" 
                                onclick="return confirmDelete('Удалить запись?')">
                                Удалить
                            </button>
                        </form>
                    </td>
                </tr>
                ''' for row in records)}
            </table>
        </div>
    ''')

@app.route('/delete/machine/<int:id>', methods=['POST'])
def delete_machine(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM machines WHERE id = ?', (id,))
        conn.commit()
    except Exception as e:
        print(f"Ошибка удаления техники: {e}")
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/machines')

@app.route('/delete/driver/<int:id>', methods=['POST'])
def delete_driver(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM drivers WHERE id = ?', (id,))
        conn.commit()
    except Exception as e:
        print(f"Ошибка удаления водителя: {e}")
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/drivers')

@app.route('/delete/counterparty/<int:id>', methods=['POST'])
def delete_counterparty(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM counterparties WHERE id = ?', (id,))
        conn.commit()
    except Exception as e:
        print(f"Ошибка удаления контрагента: {e}")
        conn.rollback()
        return "Ошибка удаления", 500
    finally:
        conn.close()
    return redirect('/admin/counterparties')

@app.route('/delete/record/<int:id>', methods=['POST'])
def delete_record(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM records WHERE id = ?', (id,))
        conn.commit()
    except Exception as e:
        print(f"Ошибка удаления записи: {e}")
        conn.rollback()
        return "Ошибка удаления записи", 500
    finally:
        conn.close()
    return redirect('/admin/records')

@app.route('/export')
def export_excel():
    conn = get_db()
    try:
        rows = conn.execute('''
            SELECT r.date, m.name, d.name, r.status,
                   r.start_time, r.end_time, r.hours,
                   c.name, r.comment
            FROM records r
            JOIN machines m ON r.machine_id = m.id
            LEFT JOIN drivers d ON r.driver_id = d.id
            LEFT JOIN counterparties c ON r.counterparty_id = c.id
            ORDER BY r.date ASC
        ''').fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "AN-30 Отчёт"

    headers = [
        "Дата", "Техника", "Водитель", "Статус",
        "Начало работы", "Конец работы", "Часы",
        "Контрагент", "Комментарий"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="444444", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 20

    for row in rows:
        date_fmt = datetime.strptime(row[0], '%Y-%m-%d').strftime('%d.%m.%Y')
        # убираем '#' у цвета, чтобы остался только HEX
        status_hex = COLORS['status'].get(row[3], '#FFFFFF')[1:]
        ws.append([
            date_fmt,
            row[1] or "-",
            row[2] or "-",
            row[3].capitalize(),
            row[4] or "-",
            row[5] or "-",
            row[6] or "0",
            row[7] or "-",
            row[8] or "-"
        ])
        # подсветка статуса
        status_cell = ws.cell(row=ws.max_row, column=4)
        status_cell.fill = PatternFill(start_color=status_hex, fill_type="solid")

    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
